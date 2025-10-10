from flask import Blueprint, request, jsonify, render_template, flash, redirect, url_for, session
import sqlite3
import os
from utils.db import get_db_connection
from werkzeug.utils import secure_filename
from routes.login import login_required
from datetime import datetime, timedelta

# Configuración del Blueprint
admin_bp = Blueprint('admin', __name__, template_folder='templates')

# Configuración de subida de imagenes
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def ensure_upload_folder():
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)

def is_admin():
    return session.get('rol') == 'admin'

def crear_notificacion(tipo, titulo, mensaje, fk_usuario=None, fk_prestamo=None):
    """Crea una notificación en el sistema"""
    conn = get_db_connection()
    try:
        conn.execute('''
            INSERT INTO notificaciones (tipo, titulo, mensaje, fk_usuario, fk_prestamo)
            VALUES (?, ?, ?, ?, ?)
        ''', (tipo, titulo, mensaje, fk_usuario, fk_prestamo))
        conn.commit()
    except Exception as e:
        print(f"Error al crear notificación: {e}")
    finally:
        conn.close()

# Rutas del Blueprint
@admin_bp.route('/')
@admin_bp.route('/')
@login_required
def admin():
    if not is_admin():
        flash('No tienes permisos para acceder a esta página', 'error')
        return redirect('/')
    
    conn = get_db_connection()
    
    # Obtener estadísticas
    total_implementos = conn.execute('SELECT COUNT(*) as count FROM implementos').fetchone()['count']
    total_usuarios = conn.execute('SELECT COUNT(*) as count FROM usuarios').fetchone()['count']
    total_prestamos = conn.execute('SELECT COUNT(*) as count FROM prestamos').fetchone()['count']
    prestamos_activos = conn.execute('SELECT COUNT(*) as count FROM prestamos WHERE fecha_devolucion IS NULL').fetchone()['count']
    
    # Obtener implementos recientes
    implementos = conn.execute('SELECT * FROM implementos ORDER BY fecha_creacion DESC LIMIT 5').fetchall()
    
    # Obtener notificaciones no leídas
    notificaciones = conn.execute('''
        SELECT n.*, u.nombre as usuario_nombre
        FROM notificaciones n
        LEFT JOIN usuarios u ON n.fk_usuario = u.id
        WHERE n.leida = 0
        ORDER BY n.fecha_creacion DESC
        LIMIT 10
    ''').fetchall()
    
    # Obtener usuarios pendientes de aprobación
    usuarios_pendientes = conn.execute('''
        SELECT id, nombre, email, rol, fecha_registro
        FROM usuarios 
        WHERE activo = 0 
        ORDER BY fecha_registro DESC
        LIMIT 5
    ''').fetchall()
    
    conn.close()
    
    return render_template('admin/panel_administrador.html',
                    total_implementos=total_implementos,
                    total_usuarios=total_usuarios,
                    total_prestamos=total_prestamos,
                    prestamos_activos=prestamos_activos,
                    implementos=implementos,
                    notificaciones=notificaciones,
                    usuarios_pendientes=usuarios_pendientes)

@admin_bp.route('/catalogo')
@login_required
def ver_catalogo():
    if not is_admin():
        flash('No tienes permisos para acceder a esta página', 'error')
        return redirect('/')
    conn = get_db_connection()
    implementos = conn.execute('SELECT * FROM implementos ORDER BY id DESC').fetchall()
    conn.close()
    return render_template('admin/panel_administrador.html', implementos=implementos)

@admin_bp.route('/catalogo/agregar', methods=['GET', 'POST'])
@login_required
def agregar_implemento():
    if not is_admin():
        flash('No tienes permisos para acceder a esta página', 'error')
        return redirect('/')
    
    if request.method == 'POST':
        implemento = request.form.get('implemento')
        descripcion = request.form.get('descripcion')
        disponibilidad = request.form.get('disponibilidad')
        categoria = request.form.get('categoria')
        
        if not all([implemento, descripcion, disponibilidad, categoria]):
            flash('Todos los campos son obligatorios', 'error')
            return render_template('admin/agregar_implemento.html')
        
        imagen_url = None
        if 'imagen' in request.files:
            file = request.files['imagen']
            if file and file.filename != '' and allowed_file(file.filename):
                ensure_upload_folder()
                filename = secure_filename(file.filename)
                unique_filename = f"{os.urandom(8).hex()}_{filename}"
                file.save(os.path.join(UPLOAD_FOLDER, unique_filename))
                imagen_url = unique_filename
        
        conn = get_db_connection()
        try:
            conn.execute(
                'INSERT INTO implementos (implemento, descripcion, disponibilidad, categoria, imagen_url) VALUES (?, ?, ?, ?, ?)',
                (implemento, descripcion, disponibilidad, categoria, imagen_url)
            )
            conn.commit()
            
            # Crear notificación
            crear_notificacion(
                'implemento_nuevo',
                'Nuevo implemento agregado',
                f'{implemento} ha sido agregado al sistema por {session.get("user_nombre")}',
                session.get('user_id')
            )
            
            flash('Implemento agregado correctamente', 'success')
        except sqlite3.Error as e:
            flash(f'Error al guardar en la base de datos: {str(e)}', 'error')
        finally:
            conn.close()
        
        return redirect(url_for('admin.ver_catalogo'))
    
    return render_template('admin/agregar_implemento.html')

@admin_bp.route('/catalogo/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_implemento(id):
    if not is_admin():
        flash('No tienes permisos para acceder a esta página', 'error')
        return redirect('/')
    
    conn = get_db_connection()
    
    if request.method == 'POST':
        implemento = request.form.get('implemento')
        descripcion = request.form.get('descripcion')
        disponibilidad = request.form.get('disponibilidad')
        categoria = request.form.get('categoria')
        estado = request.form.get('estado', 'Bueno')
        
        if not all([implemento, descripcion, disponibilidad, categoria]):
            flash('Todos los campos son obligatorios', 'error')
            return redirect(url_for('admin.editar_implemento', id=id))
        
        imagen_url = None
        if 'imagen' in request.files:
            file = request.files['imagen']
            if file and file.filename != '' and allowed_file(file.filename):
                ensure_upload_folder()
                filename = secure_filename(file.filename)
                unique_filename = f"{os.urandom(8).hex()}_{filename}"
                file.save(os.path.join(UPLOAD_FOLDER, unique_filename))
                imagen_url = unique_filename
        
        try:
            if imagen_url:
                conn.execute(
                    'UPDATE implementos SET implemento = ?, descripcion = ?, disponibilidad = ?, categoria = ?, imagen_url = ?, estado = ?, fecha_actualizacion = CURRENT_TIMESTAMP WHERE id = ?',
                    (implemento, descripcion, disponibilidad, categoria, imagen_url, estado, id)
                )
            else:
                conn.execute(
                    'UPDATE implementos SET implemento = ?, descripcion = ?, disponibilidad = ?, categoria = ?, estado = ?, fecha_actualizacion = CURRENT_TIMESTAMP WHERE id = ?',
                    (implemento, descripcion, disponibilidad, categoria, estado, id)
                )
            conn.commit()
            flash('Implemento actualizado correctamente', 'success')
        except sqlite3.Error as e:
            flash(f'Error al actualizar: {str(e)}', 'error')
        
        conn.close()
        return redirect(url_for('admin.ver_catalogo'))
    
    implemento = conn.execute('SELECT * FROM implementos WHERE id = ?', (id,)).fetchone()
    conn.close()
    
    if implemento is None:
        flash('Implemento no encontrado', 'error')
        return redirect(url_for('admin.ver_catalogo'))
    
    return render_template('admin/editar_implemento.html', implemento=implemento)

@admin_bp.route('/catalogo/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar_implemento(id):
    if not is_admin():
        flash('No tienes permisos para acceder a esta página', 'error')
        return redirect('/')
    
    conn = get_db_connection()
    
    try:
        implemento = conn.execute('SELECT imagen_url FROM implementos WHERE id = ?', (id,)).fetchone()
        
        if implemento and implemento['imagen_url']:
            try:
                os.remove(os.path.join(UPLOAD_FOLDER, implemento['imagen_url']))
            except OSError:
                pass
        
        conn.execute('DELETE FROM implementos WHERE id = ?', (id,))
        conn.commit()
        flash('Implemento eliminado correctamente', 'success')
    except sqlite3.Error as e:
        flash(f'Error al eliminar: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('admin.ver_catalogo'))

# Gestión de usuarios
@admin_bp.route('/usuarios')
@login_required
def gestion_usuarios():
    if not is_admin():
        flash('No tienes permisos para acceder a esta página', 'error')
        return redirect('/')
    
    conn = get_db_connection()
    try:
        # Obtener y validar filtros
        filtro_estado = request.args.get('estado', 'todos')
        filtro_rol = request.args.get('rol', 'todos')
        
        # Validar valores de filtro
        estados_validos = ['todos', 'activos', 'pendientes']
        roles_validos = ['todos', 'admin', 'instructor', 'funcionario']
        
        if filtro_estado not in estados_validos:
            filtro_estado = 'todos'
        
        if filtro_rol not in roles_validos:
            filtro_rol = 'todos'
        
        # Construir query base
        query = "SELECT * FROM usuarios WHERE 1=1"
        params = []
        
        # Aplicar filtros
        if filtro_estado == 'activos':
            query += " AND activo = 1"
        elif filtro_estado == 'pendientes':
            query += " AND activo = 0"
        
        if filtro_rol != 'todos':
            query += " AND rol = ?"
            params.append(filtro_rol)
        
        query += " ORDER BY id DESC"
        
        usuarios = conn.execute(query, params).fetchall()
        
        conn.close()
        
        return render_template('admin/gestion_usuarios.html', 
                             usuarios=usuarios,
                             filtro_estado=filtro_estado,
                             filtro_rol=filtro_rol)
    except Exception as e:
        flash(f'Error al cargar usuarios: {str(e)}', 'error')
        conn.close()
        return redirect('/admin')

# Gestión de préstamos - Devoluciones
@admin_bp.route('/devolucion_prestamos')
@login_required
def devolucion_prestamos():
    if not is_admin():
        flash('No tienes permisos para acceder a esta página', 'error')
        return redirect('/')

    conn = get_db_connection()
    
    try:
        prestamos_activos = conn.execute('''
            SELECT p.*, u.nombre as usuario_nombre, i.implemento,
                   julianday('now') - julianday(p.fecha_prestamo) as dias_transcurridos
            FROM prestamos p
            JOIN usuarios u ON p.fk_usuario = u.id
            JOIN implementos i ON p.fk_implemento = i.id
            WHERE p.fecha_devolucion IS NULL
            ORDER BY p.fecha_prestamo DESC
        ''').fetchall()
        
        prestamos_con_dias = []
        for prestamo in prestamos_activos:
            prestamo_dict = dict(prestamo)
            prestamo_dict['dias_transcurridos'] = int(prestamo['dias_transcurridos']) if prestamo['dias_transcurridos'] else 0
            prestamos_con_dias.append(prestamo_dict)
        
        total_prestamos = len(prestamos_con_dias)
        implementos_unicos = len(set(p['fk_implemento'] for p in prestamos_con_dias))
        
        hoy = datetime.now().strftime("%Y-%m-%d")
        prestamos_hoy = conn.execute('''
            SELECT COUNT(*) as count FROM prestamos 
            WHERE DATE(fecha_prestamo) = ? AND fecha_devolucion IS NULL
        ''', (hoy,)).fetchone()['count']
        
    except Exception as e:
        flash(f'Error al cargar préstamos: {str(e)}', 'error')
        prestamos_con_dias = []
        total_prestamos = 0
        implementos_unicos = 0
        prestamos_hoy = 0
    finally:
        conn.close()
    
    return render_template('admin/dvprestamos.html',
                         prestamos_activos=prestamos_con_dias,
                         total_prestamos=total_prestamos,
                         total_implementos=implementos_unicos,
                         prestamos_hoy=prestamos_hoy)

# Procesar devolución desde el panel admin
@admin_bp.route('/devolver_prestamo_admin/<int:id>', methods=['POST'])
@login_required
def devolver_prestamo_admin(id):
    print(f"DEBUG: Procesando devolución para préstamo ID: {id}")
    print(f"DEBUG: Usuario actual: {session.get('user_id')}, Rol: {session.get('rol')}")
    
    if not is_admin():
        flash('No tienes permisos para acceder a esta página', 'error')
        return redirect('/')
    
    conn = get_db_connection()
    try:
        # Obtener datos del formulario
        novedad = request.form.get('novedad', 'Ninguna')
        estado_implemento = request.form.get('estado_implemento', 'Bueno')
        observaciones = request.form.get('observaciones', '')
        
        print(f"DEBUG: Datos del formulario - Novedad: {novedad}, Estado: {estado_implemento}, Observaciones: {observaciones}")
        
        prestamo = conn.execute('''
            SELECT p.*, i.implemento, i.disponibilidad, u.nombre as usuario_nombre
            FROM prestamos p
            JOIN implementos i ON p.fk_implemento = i.id
            JOIN usuarios u ON p.fk_usuario = u.id
            WHERE p.id = ?
        ''', (id,)).fetchone()
        
        if not prestamo:
            flash('No se encontró el préstamo.', 'error')
            return redirect(url_for('admin.devolucion_prestamos'))
        
        if prestamo['fecha_devolucion'] is not None:
            flash('Este préstamo ya fue devuelto anteriormente.', 'warning')
            return redirect(url_for('admin.devolucion_prestamos'))
        
        # Registrar la devolución
        fecha_devolucion = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn.execute('''
            UPDATE prestamos 
            SET fecha_devolucion = ?, novedad = ?, estado_implemento_devolucion = ?, observaciones = ?
            WHERE id = ?
        ''', (fecha_devolucion, novedad, estado_implemento, observaciones, id))
        
        # Actualizar disponibilidad del implemento según la novedad
        novedades_que_reducen_cantidad = ['Daño', 'Robo', 'Desgaste excesivo', 'Pérdida']
        cantidad_a_reducir = 0
        
        if novedad in novedades_que_reducen_cantidad:
            cantidad_a_reducir = 1
            print(f"DEBUG: Novedad '{novedad}' detectada - reduciendo cantidad en {cantidad_a_reducir}")
        
        # Calcular nueva disponibilidad (normalmente +1, pero -1 si hay novedad grave)
        nueva_disponibilidad = prestamo['disponibilidad'] + 1 - cantidad_a_reducir
        
        # Asegurar que la disponibilidad no sea negativa
        if nueva_disponibilidad < 0:
            nueva_disponibilidad = 0
        
        conn.execute(
            'UPDATE implementos SET disponibilidad = ? WHERE id = ?',
            (nueva_disponibilidad, prestamo['fk_implemento'])
        )
        
        # Actualizar estado del implemento si es necesario
        if estado_implemento != 'Bueno':
            conn.execute(
                'UPDATE implementos SET estado = ? WHERE id = ?',
                (estado_implemento, prestamo['fk_implemento'])
            )
        
        conn.commit()
        
        # Crear notificación con información sobre la novedad
        mensaje_notif = f'Devolución de {prestamo["implemento"]} por {prestamo["usuario_nombre"]}'
        if novedad != 'Ninguna':
            mensaje_notif += f' - Novedad: {novedad}'
            if novedad in novedades_que_reducen_cantidad:
                mensaje_notif += ' - Se redujo la cantidad disponible del implemento'
        
        crear_notificacion(
            'devolucion',
            'Devolución registrada',
            mensaje_notif,
            session.get('user_id'),
            id
        )
        
        flash(f'Devolución registrada exitosamente: {prestamo["implemento"]}', 'success')
        
    except Exception as e:
        flash(f'Error al procesar la devolución: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('admin.devolucion_prestamos'))

# API para notificaciones
@admin_bp.route('/api/notificaciones')
@login_required
def api_notificaciones():
    if not is_admin():
        return jsonify({'error': 'Sin permisos'}), 403
    
    conn = get_db_connection()
    notificaciones = conn.execute('''
        SELECT n.*, u.nombre as usuario_nombre
        FROM notificaciones n
        LEFT JOIN usuarios u ON n.fk_usuario = u.id
        WHERE n.leida = 0
        ORDER BY n.fecha_creacion DESC
        LIMIT 20
    ''').fetchall()
    conn.close()
    
    return jsonify([dict(notif) for notif in notificaciones])

# Marcar notificación como leída
@admin_bp.route('/api/notificaciones/<int:id>/leer', methods=['POST'])
@login_required
def marcar_notificacion_leida(id):
    print(f"DEBUG: Intentando marcar notificación {id} como leída")
    print(f"DEBUG: Usuario actual: {session.get('user_id')}, Rol: {session.get('rol')}")
    
    if not is_admin():
        print("DEBUG: Usuario no tiene permisos de admin")
        return jsonify({'success': False, 'error': 'Sin permisos'}), 403
    
    conn = get_db_connection()
    try:
        # Verificar que la notificación existe
        notificacion = conn.execute('SELECT * FROM notificaciones WHERE id = ?', (id,)).fetchone()
        if not notificacion:
            print(f"DEBUG: Notificación {id} no encontrada")
            return jsonify({'success': False, 'error': 'Notificación no encontrada'}), 404
        
        print(f"DEBUG: Notificación encontrada: {dict(notificacion)}")
        
        # Verificar si ya está marcada como leída
        if notificacion['leida'] == 1:
            print(f"DEBUG: Notificación {id} ya está marcada como leída")
            return jsonify({'success': True, 'message': 'Notificación ya estaba marcada como leída'})
        
        # Marcar como leída
        conn.execute('UPDATE notificaciones SET leida = 1 WHERE id = ?', (id,))
        conn.commit()
        print(f"DEBUG: Notificación {id} marcada como leída exitosamente")
        return jsonify({'success': True, 'message': 'Notificación marcada como leída'})
    except Exception as e:
        print(f"ERROR: Error al marcar notificación como leída: {e}")
        return jsonify({'success': False, 'error': f'Error del servidor: {str(e)}'}), 500
    finally:
        conn.close()

# Marcar todas las notificaciones como leídas
@admin_bp.route('/api/notificaciones/leer_todas', methods=['POST'])
@login_required
def marcar_todas_leidas():
    if not is_admin():
        return jsonify({'error': 'Sin permisos'}), 403
    
    conn = get_db_connection()
    try:
        # Contar cuántas notificaciones se van a marcar
        cursor = conn.execute('SELECT COUNT(*) as count FROM notificaciones WHERE leida = 0')
        count = cursor.fetchone()['count']
        
        # Marcar todas como leídas
        conn.execute('UPDATE notificaciones SET leida = 1 WHERE leida = 0')
        conn.commit()
        print(f"{count} notificaciones marcadas como leídas")
        return jsonify({'success': True, 'message': f'{count} notificaciones marcadas como leídas'})
    except Exception as e:
        print(f"Error al marcar todas las notificaciones como leídas: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

# Página de notificaciones
@admin_bp.route('/notificaciones')
@login_required
def notificaciones():
    if not is_admin():
        flash('No tienes permisos para acceder a esta página', 'error')
        return redirect('/')
    
    conn = get_db_connection()
    try:
        # Obtener todas las notificaciones
        notificaciones = conn.execute('''
            SELECT n.*, u.nombre as usuario_nombre
            FROM notificaciones n
            LEFT JOIN usuarios u ON n.fk_usuario = u.id
            ORDER BY n.fecha_creacion DESC
            LIMIT 50
        ''').fetchall()
        
        conn.close()
        
        return render_template('admin/notificaciones.html', notificaciones=notificaciones)
    except Exception as e:
        flash(f'Error al cargar notificaciones: {str(e)}', 'error')
        conn.close()
        return redirect('/admin')

# Eliminar notificación individual
@admin_bp.route('/api/notificaciones/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar_notificacion(id):
    if not is_admin():
        return jsonify({'error': 'Sin permisos'}), 403
    
    conn = get_db_connection()
    try:
        conn.execute('DELETE FROM notificaciones WHERE id = ?', (id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

# Eliminar todas las notificaciones leídas
@admin_bp.route('/api/notificaciones/eliminar_leidas', methods=['POST'])
@login_required
def eliminar_notificaciones_leidas():
    if not is_admin():
        return jsonify({'error': 'Sin permisos'}), 403
    
    conn = get_db_connection()
    try:
        conn.execute('DELETE FROM notificaciones WHERE leida = 1')
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

# Gestión de préstamos - Vista principal
@admin_bp.route('/gestion_prestamos')
@login_required
def gestion_prestamos():
    conn = get_db_connection()
    try:
        # Obtener implementos disponibles
        implementos_disponibles = conn.execute('''
            SELECT id, implemento, disponibilidad, categoria 
            FROM implementos 
            WHERE disponibilidad > 0 
            ORDER BY implemento
        ''').fetchall()
        
        # Obtener y validar filtros
        filtro_estado = request.args.get('estado', 'todos')
        filtro_dias = request.args.get('dias', '30')
        
        # Validar valores de filtro
        estados_validos = ['todos', 'activos', 'devueltos']
        if filtro_estado not in estados_validos:
            filtro_estado = 'todos'
        
        # Validar días
        try:
            filtro_dias = int(filtro_dias)
            if filtro_dias < 0:
                filtro_dias = 30
        except (ValueError, TypeError):
            filtro_dias = 30
        
        # Construir query base para préstamos
        query = '''
            SELECT p.*, u.nombre as usuario, i.implemento,
                   julianday('now') - julianday(p.fecha_prestamo) as dias_transcurridos
            FROM prestamos p
            JOIN usuarios u ON p.fk_usuario = u.id
            JOIN implementos i ON p.fk_implemento = i.id
            WHERE 1=1
        '''
        params = []
        
        # Aplicar filtros
        if filtro_estado == 'activos':
            query += " AND p.fecha_devolucion IS NULL"
        elif filtro_estado == 'devueltos':
            query += " AND p.fecha_devolucion IS NOT NULL"
        
        # Filtro por días
        if filtro_dias > 0:
            fecha_limite = (datetime.now() - timedelta(days=filtro_dias)).strftime("%Y-%m-%d")
            query += " AND DATE(p.fecha_prestamo) >= ?"
            params.append(fecha_limite)
        
        query += " ORDER BY p.fecha_prestamo DESC"
        
        prestamos = conn.execute(query, params).fetchall()
        
    except Exception as e:
        flash(f'Error al cargar datos: {str(e)}', 'error')
        implementos_disponibles = []
        prestamos = []
    finally:
        conn.close()
    
    return render_template('views/gestion_prestamos.html',
                         implementos_disponibles=implementos_disponibles,
                         prestamos=prestamos,
                         filtro_estado=filtro_estado,
                         filtro_dias=filtro_dias)

# Registrar préstamo individual
@admin_bp.route('/registrar_prestamo_individual', methods=['POST'])
@login_required
def registrar_prestamo_individual():
    # Solo instructores, funcionarios y administradores pueden hacer préstamos
    if session.get('rol') not in ['instructor', 'funcionario', 'admin']:
        flash('No tienes permiso para realizar préstamos.', 'error')
        return redirect(url_for('admin.gestion_prestamos'))
    
    conn = get_db_connection()
    try:
        implemento_id = request.form.get('implemento_id')
        nombre_prestatario = request.form.get('nombre_prestatario')
        instructor = request.form.get('instructor')
        jornada = request.form.get('jornada')
        ambiente = request.form.get('ambiente')
        
        # Verificar disponibilidad del implemento
        implemento = conn.execute(
            'SELECT id, implemento, disponibilidad FROM implementos WHERE id = ?', (implemento_id,)
        ).fetchone()

        if not implemento:
            flash('El implemento no existe.', 'error')
            return redirect(url_for('admin.gestion_prestamos'))

        if implemento['disponibilidad'] <= 0:
            flash('Este implemento no está disponible para préstamo.', 'error')
            return redirect(url_for('admin.gestion_prestamos'))

        fk_usuario = session.get('user_id')
        fecha_prestamo = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Registrar el préstamo individual
        cursor = conn.execute('''
            INSERT INTO prestamos (fk_usuario, fk_implemento, tipo_prestamo, nombre_prestatario, 
                                instructor, jornada, ambiente, fecha_prestamo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (fk_usuario, implemento_id, 'individual', nombre_prestatario, instructor, jornada, ambiente, fecha_prestamo))
        
        prestamo_id = cursor.lastrowid

        # Actualizar la disponibilidad del implemento
        nueva_disponibilidad = implemento['disponibilidad'] - 1
        conn.execute('UPDATE implementos SET disponibilidad = ? WHERE id = ?', 
                    (nueva_disponibilidad, implemento_id))

        conn.commit()
        
        # Crear notificación para admin
        crear_notificacion(
            'prestamo_individual',
            'Nuevo préstamo individual',
            f'{nombre_prestatario} ha solicitado un préstamo de {implemento["implemento"]}',
            fk_usuario,
            prestamo_id
        )
        
        flash(f"Préstamo de '{implemento['implemento']}' registrado con éxito", "success")
        
    except Exception as e:
        flash(f"Error en el préstamo: {str(e)}", "error")
    finally:
        conn.close()

    return redirect(url_for('admin.gestion_prestamos'))

# Registrar préstamo múltiple
@admin_bp.route('/registrar_prestamo_multiple', methods=['POST'])
@login_required
def registrar_prestamo_multiple():
    # Solo instructores, funcionarios y administradores pueden hacer préstamos
    if session.get('rol') not in ['instructor', 'funcionario', 'admin']:
        flash('No tienes permiso para realizar préstamos.', 'error')
        return redirect(url_for('admin.gestion_prestamos'))
    
    conn = get_db_connection()
    try:
        implemento_id = request.form.get('implemento_id')
        ficha = request.form.get('ficha')
        ambiente = request.form.get('ambiente')
        horario = request.form.get('horario')
        
        # El nombre del prestatario es el usuario que hace el préstamo
        nombre_prestatario = session.get('user_nombre')
        
        # Verificar disponibilidad del implemento
        implemento = conn.execute(
            'SELECT id, implemento, disponibilidad FROM implementos WHERE id = ?', (implemento_id,)
        ).fetchone()

        if not implemento:
            flash('El implemento no existe.', 'error')
            return redirect(url_for('admin.gestion_prestamos'))

        if implemento['disponibilidad'] <= 0:
            flash('Este implemento no está disponible para préstamo.', 'error')
            return redirect(url_for('admin.gestion_prestamos'))

        fk_usuario = session.get('user_id')
        fecha_prestamo = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Registrar el préstamo múltiple
        cursor = conn.execute('''
            INSERT INTO prestamos (fk_usuario, fk_implemento, tipo_prestamo, nombre_prestatario, 
                                ficha, ambiente, horario, fecha_prestamo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (fk_usuario, implemento_id, 'multiple', nombre_prestatario, ficha, ambiente, horario, fecha_prestamo))
        
        prestamo_id = cursor.lastrowid

        # Actualizar la disponibilidad del implemento
        nueva_disponibilidad = implemento['disponibilidad'] - 1
        conn.execute('UPDATE implementos SET disponibilidad = ? WHERE id = ?', 
                    (nueva_disponibilidad, implemento_id))

        conn.commit()
        
        # Crear notificación para admin
        crear_notificacion(
            'prestamo_multiple',
            'Nuevo préstamo múltiple',
            f'{nombre_prestatario} ha solicitado un préstamo múltiple de {implemento["implemento"]} para ficha {ficha}',
            fk_usuario,
            prestamo_id
        )
        
        flash(f"Préstamo múltiple de '{implemento['implemento']}' registrado con éxito", "success")
        
    except Exception as e:
        flash(f"Error en el préstamo: {str(e)}", "error")
    finally:
        conn.close()

    return redirect(url_for('admin.gestion_prestamos'))

# Procesar devolución de préstamo (solo admin)
@admin_bp.route('/devolver_prestamo/<int:id>', methods=['POST'])
@login_required
def devolver_prestamo(id):
    if not is_admin():
        flash('No tienes permiso para procesar devoluciones.', 'error')
        return redirect(url_for('admin.gestion_prestamos'))

    conn = get_db_connection()
    try:
        # Obtener datos del formulario
        novedad = request.form.get('novedad', 'Ninguna')
        estado_implemento = request.form.get('estado_implemento', 'Bueno')
        observaciones = request.form.get('observaciones', '')
        
        # Obtener información completa del préstamo
        prestamo = conn.execute('''
            SELECT p.*, i.implemento, i.disponibilidad, u.nombre as usuario_nombre
            FROM prestamos p
            JOIN implementos i ON p.fk_implemento = i.id
            JOIN usuarios u ON p.fk_usuario = u.id
            WHERE p.id = ?
        ''', (id,)).fetchone()

        if not prestamo:
            flash('No se encontró el préstamo.', 'error')
            return redirect(url_for('admin.gestion_prestamos'))

        if prestamo['fecha_devolucion'] is not None:
            flash('Este préstamo ya fue devuelto anteriormente.', 'warning')
            return redirect(url_for('admin.gestion_prestamos'))

        # Registrar la devolución
        fecha_devolucion = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn.execute('''
            UPDATE prestamos 
            SET fecha_devolucion = ?, novedad = ?, estado_implemento_devolucion = ?, observaciones = ?
            WHERE id = ?
        ''', (fecha_devolucion, novedad, estado_implemento, observaciones, id))

        # Actualizar disponibilidad del implemento según la novedad
        novedades_que_reducen_cantidad = ['Daño', 'Robo', 'Desgaste excesivo', 'Pérdida']
        cantidad_a_reducir = 0
        
        if novedad in novedades_que_reducen_cantidad:
            cantidad_a_reducir = 1
            print(f"DEBUG: Novedad '{novedad}' detectada - reduciendo cantidad en {cantidad_a_reducir}")
        
        # Calcular nueva disponibilidad (normalmente +1, pero -1 si hay novedad grave)
        nueva_disponibilidad = prestamo['disponibilidad'] + 1 - cantidad_a_reducir
        
        # Asegurar que la disponibilidad no sea negativa
        if nueva_disponibilidad < 0:
            nueva_disponibilidad = 0
        
        conn.execute(
            'UPDATE implementos SET disponibilidad = ? WHERE id = ?',
            (nueva_disponibilidad, prestamo['fk_implemento'])
        )
        
        # Actualizar estado del implemento si es necesario
        if estado_implemento != 'Bueno':
            conn.execute(
                'UPDATE implementos SET estado = ? WHERE id = ?',
                (estado_implemento, prestamo['fk_implemento'])
            )

        conn.commit()
        
        # Crear notificación con información sobre la novedad
        mensaje_notif = f'Devolución de {prestamo["implemento"]} por {prestamo["usuario_nombre"]}'
        if novedad != 'Ninguna':
            mensaje_notif += f' - Novedad: {novedad}'
            if novedad in novedades_que_reducen_cantidad:
                mensaje_notif += ' - Se redujo la cantidad disponible del implemento'
        if estado_implemento != 'Bueno':
            mensaje_notif += f' - Estado: {estado_implemento}'
        
        crear_notificacion(
            'devolucion',
            'Devolución registrada',
            mensaje_notif,
            session.get('user_id'),
            id
        )
        
        flash(f'Devolución registrada exitosamente: {prestamo["implemento"]}', 'success')
    except Exception as e:
        flash(f'Error al procesar la devolución: {str(e)}', 'error')
    finally:
        conn.close()

    return redirect(url_for('admin.gestion_prestamos'))

# API estadísticas
@admin_bp.route('/api/admin/estadisticas')
def api_estadisticas():
    conn = get_db_connection()
    
    total_implementos = conn.execute('SELECT COUNT(*) as count FROM implementos').fetchone()['count']
    total_usuarios = conn.execute('SELECT COUNT(*) as count FROM usuarios').fetchone()['count']
    total_prestamos = conn.execute('SELECT COUNT(*) as count FROM prestamos').fetchone()['count']
    prestamos_activos = conn.execute('SELECT COUNT(*) as count FROM prestamos WHERE fecha_devolucion IS NULL').fetchone()['count']
    
    conn.close()
    
    return jsonify({
        'total_implementos': total_implementos,
        'total_usuarios': total_usuarios,
        'total_prestamos': total_prestamos,
        'prestamos_activos': prestamos_activos
    })

# Gestión de préstamos para instructores
@admin_bp.route('/gestion_prestamos_instructores')
@login_required
def gestion_prestamos_instructores():
    # Solo instructores y funcionarios pueden acceder
    if session.get('rol') not in ['instructor', 'funcionario']:
        flash('No tienes permisos para acceder a esta página', 'error')
        return redirect('/')
    
    conn = get_db_connection()
    try:
        # Obtener y validar filtros
        filtro_estado = request.args.get('estado', 'todos')
        
        # Validar valores de filtro
        estados_validos = ['todos', 'activos', 'devueltos']
        if filtro_estado not in estados_validos:
            filtro_estado = 'todos'
        
        # Construir query base para préstamos del usuario actual
        query = '''
            SELECT p.*, i.implemento
            FROM prestamos p
            JOIN implementos i ON p.fk_implemento = i.id
            WHERE p.fk_usuario = ?
        '''
        params = [session.get('user_id')]
        
        # Aplicar filtros
        if filtro_estado == 'activos':
            query += " AND p.fecha_devolucion IS NULL"
        elif filtro_estado == 'devueltos':
            query += " AND p.fecha_devolucion IS NOT NULL"
        
        query += " ORDER BY p.fecha_prestamo DESC"
        
        prestamos = conn.execute(query, params).fetchall()
        
    except Exception as e:
        flash(f'Error al cargar datos: {str(e)}', 'error')
        prestamos = []
    finally:
        conn.close()
    
    return render_template('views/gestion_prestamos_instructores.html',
                         prestamos=prestamos,
                         filtro_estado=filtro_estado)

# Agregar novedad a préstamo (solo instructores/funcionarios para sus préstamos)
@admin_bp.route('/agregar_novedad/<int:id>', methods=['POST'])
@login_required
def agregar_novedad(id):
    # Solo instructores y funcionarios pueden agregar novedades
    if session.get('rol') not in ['instructor', 'funcionario']:
        flash('No tienes permisos para agregar novedades', 'error')
        return redirect('/')
    
    conn = get_db_connection()
    try:
        # Verificar que el préstamo pertenece al usuario actual
        prestamo = conn.execute('''
            SELECT p.*, i.implemento
            FROM prestamos p
            JOIN implementos i ON p.fk_implemento = i.id
            WHERE p.id = ? AND p.fk_usuario = ?
        ''', (id, session.get('user_id'))).fetchone()

        if not prestamo:
            flash('No se encontró el préstamo o no tienes permisos para modificarlo.', 'error')
            return redirect(url_for('admin.gestion_prestamos_instructores'))

        if prestamo['fecha_devolucion'] is not None:
            flash('No se pueden agregar novedades a préstamos ya devueltos.', 'warning')
            return redirect(url_for('admin.gestion_prestamos_instructores'))

        # Obtener datos del formulario
        novedad = request.form.get('novedad')
        descripcion = request.form.get('descripcion')
        
        if not all([novedad, descripcion]):
            flash('Todos los campos son obligatorios.', 'error')
            return redirect(url_for('admin.gestion_prestamos_instructores'))

        # Actualizar el préstamo con la novedad
        conn.execute('''
            UPDATE prestamos 
            SET novedad = ?, observaciones = ?
            WHERE id = ?
        ''', (novedad, descripcion, id))
        
        conn.commit()
        
        # Crear notificación para admin
        crear_notificacion(
            'novedad_prestamo',
            'Novedad en préstamo',
            f'{session.get("user_nombre")} agregó una novedad al préstamo de {prestamo["implemento"]}: {novedad}',
            session.get('user_id'),
            id
        )
        
        flash(f'Novedad agregada exitosamente al préstamo de {prestamo["implemento"]}', 'success')
        
    except Exception as e:
        flash(f'Error al agregar novedad: {str(e)}', 'error')
    finally:
        conn.close()

    return redirect(url_for('admin.gestion_prestamos_instructores'))

# Gestión de préstamos para administradores
@admin_bp.route('/gestion_prestamos_admin')
@login_required
def gestion_prestamos_admin():
    if not is_admin():
        flash('No tienes permisos para acceder a esta página', 'error')
        return redirect('/')
    
    conn = get_db_connection()
    try:
        # Obtener implementos disponibles
        implementos_disponibles = conn.execute('''
            SELECT id, implemento, disponibilidad, categoria 
            FROM implementos 
            WHERE disponibilidad > 0 
            ORDER BY implemento
        ''').fetchall()
        
        # Obtener y validar filtros
        filtro_estado = request.args.get('estado', 'todos')
        filtro_dias = request.args.get('dias', '30')
        
        # Validar valores de filtro
        estados_validos = ['todos', 'activos', 'devueltos']
        if filtro_estado not in estados_validos:
            filtro_estado = 'todos'
        
        # Validar días
        try:
            filtro_dias = int(filtro_dias)
            if filtro_dias < 0:
                filtro_dias = 30
        except (ValueError, TypeError):
            filtro_dias = 30
        
        # Construir query base para préstamos
        query = '''
            SELECT p.*, u.nombre as usuario, i.implemento,
                   julianday('now') - julianday(p.fecha_prestamo) as dias_transcurridos
            FROM prestamos p
            JOIN usuarios u ON p.fk_usuario = u.id
            JOIN implementos i ON p.fk_implemento = i.id
            WHERE 1=1
        '''
        params = []
        
        # Aplicar filtros
        if filtro_estado == 'activos':
            query += " AND p.fecha_devolucion IS NULL"
        elif filtro_estado == 'devueltos':
            query += " AND p.fecha_devolucion IS NOT NULL"
        
        # Filtro por días
        if filtro_dias > 0:
            fecha_limite = (datetime.now() - timedelta(days=filtro_dias)).strftime("%Y-%m-%d")
            query += " AND DATE(p.fecha_prestamo) >= ?"
            params.append(fecha_limite)
        
        query += " ORDER BY p.fecha_prestamo DESC"
        
        prestamos = conn.execute(query, params).fetchall()
        
    except Exception as e:
        flash(f'Error al cargar datos: {str(e)}', 'error')
        implementos_disponibles = []
        prestamos = []
    finally:
        conn.close()
    
    return render_template('admin/gestion_prestamos_admin.html',
                         implementos_disponibles=implementos_disponibles,
                         prestamos=prestamos,
                         filtro_estado=filtro_estado,
                         filtro_dias=filtro_dias)

# Registrar préstamo individual como admin
@admin_bp.route('/registrar_prestamo_admin_individual', methods=['POST'])
@login_required
def registrar_prestamo_admin_individual():
    if not is_admin():
        flash('No tienes permisos para realizar esta acción', 'error')
        return redirect('/')
    
    conn = get_db_connection()
    try:
        implemento_id = request.form.get('implemento_id')
        nombre_prestatario = request.form.get('nombre_prestatario')
        instructor = request.form.get('instructor')
        jornada = request.form.get('jornada')
        ambiente = request.form.get('ambiente')
        
        # Verificar disponibilidad del implemento
        implemento = conn.execute(
            'SELECT id, implemento, disponibilidad FROM implementos WHERE id = ?', (implemento_id,)
        ).fetchone()

        if not implemento:
            flash('El implemento no existe.', 'error')
            return redirect(url_for('admin.gestion_prestamos_admin'))

        if implemento['disponibilidad'] <= 0:
            flash('Este implemento no está disponible para préstamo.', 'error')
            return redirect(url_for('admin.gestion_prestamos_admin'))

        fk_usuario = session.get('user_id')
        fecha_prestamo = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Registrar el préstamo individual
        cursor = conn.execute('''
            INSERT INTO prestamos (fk_usuario, fk_implemento, tipo_prestamo, nombre_prestatario, 
                                instructor, jornada, ambiente, fecha_prestamo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (fk_usuario, implemento_id, 'individual', nombre_prestatario, instructor, jornada, ambiente, fecha_prestamo))
        
        prestamo_id = cursor.lastrowid

        # Actualizar la disponibilidad del implemento
        nueva_disponibilidad = implemento['disponibilidad'] - 1
        conn.execute('UPDATE implementos SET disponibilidad = ? WHERE id = ?', 
                    (nueva_disponibilidad, implemento_id))

        conn.commit()
        
        # Crear notificación
        crear_notificacion(
            'prestamo_admin',
            'Préstamo registrado por admin',
            f'Admin registró préstamo individual de {implemento["implemento"]} para {nombre_prestatario}',
            fk_usuario,
            prestamo_id
        )
        
        flash(f"Préstamo de '{implemento['implemento']}' registrado con éxito", "success")
        
    except Exception as e:
        flash(f"Error en el préstamo: {str(e)}", "error")
    finally:
        conn.close()

    return redirect(url_for('admin.gestion_prestamos_admin'))

# Registrar préstamo múltiple como admin
@admin_bp.route('/registrar_prestamo_admin_multiple', methods=['POST'])
@login_required
def registrar_prestamo_admin_multiple():
    if not is_admin():
        flash('No tienes permisos para realizar esta acción', 'error')
        return redirect('/')
    
    conn = get_db_connection()
    try:
        implemento_id = request.form.get('implemento_id')
        ficha = request.form.get('ficha')
        instructor = request.form.get('instructor')
        ambiente = request.form.get('ambiente')
        horario = request.form.get('horario')
        
        # El nombre del prestatario es el usuario que hace el préstamo
        nombre_prestatario = session.get('user_nombre')
        
        # Verificar disponibilidad del implemento
        implemento = conn.execute(
            'SELECT id, implemento, disponibilidad FROM implementos WHERE id = ?', (implemento_id,)
        ).fetchone()

        if not implemento:
            flash('El implemento no existe.', 'error')
            return redirect(url_for('admin.gestion_prestamos_admin'))

        if implemento['disponibilidad'] <= 0:
            flash('Este implemento no está disponible para préstamo.', 'error')
            return redirect(url_for('admin.gestion_prestamos_admin'))

        fk_usuario = session.get('user_id')
        fecha_prestamo = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Registrar el préstamo múltiple
        cursor = conn.execute('''
            INSERT INTO prestamos (fk_usuario, fk_implemento, tipo_prestamo, nombre_prestatario, 
                                instructor, jornada, ficha, ambiente, horario, fecha_prestamo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (fk_usuario, implemento_id, 'multiple', nombre_prestatario, instructor, 'N/A', ficha, ambiente, horario, fecha_prestamo))
        
        prestamo_id = cursor.lastrowid

        # Actualizar la disponibilidad del implemento
        nueva_disponibilidad = implemento['disponibilidad'] - 1
        conn.execute('UPDATE implementos SET disponibilidad = ? WHERE id = ?', 
                    (nueva_disponibilidad, implemento_id))

        conn.commit()
        
        # Crear notificación
        crear_notificacion(
            'prestamo_admin',
            'Préstamo múltiple registrado por admin',
            f'Admin registró préstamo múltiple de {implemento["implemento"]} para ficha {ficha}',
            fk_usuario,
            prestamo_id
        )
        
        flash(f"Préstamo múltiple de '{implemento['implemento']}' registrado con éxito", "success")
        
    except Exception as e:
        flash(f"Error en el préstamo: {str(e)}", "error")
    finally:
        conn.close()

    return redirect(url_for('admin.gestion_prestamos_admin'))

# Editar préstamo (solo admin)
@admin_bp.route('/editar_prestamo/<int:id>', methods=['POST'])
@login_required
def editar_prestamo(id):
    if not is_admin():
        flash('No tienes permisos para editar préstamos', 'error')
        return redirect('/')
    
    conn = get_db_connection()
    try:
        # Obtener datos del formulario
        instructor = request.form.get('instructor')
        jornada = request.form.get('jornada')
        ambiente = request.form.get('ambiente')
        
        if not all([instructor, jornada, ambiente]):
            flash('Todos los campos son obligatorios.', 'error')
            return redirect(url_for('admin.gestion_prestamos_admin'))
        
        # Verificar que el préstamo existe y está activo
        prestamo = conn.execute('''
            SELECT p.*, i.implemento
            FROM prestamos p
            JOIN implementos i ON p.fk_implemento = i.id
            WHERE p.id = ?
        ''', (id,)).fetchone()

        if not prestamo:
            flash('No se encontró el préstamo.', 'error')
            return redirect(url_for('admin.gestion_prestamos_admin'))

        if prestamo['fecha_devolucion'] is not None:
            flash('No se pueden editar préstamos ya devueltos.', 'warning')
            return redirect(url_for('admin.gestion_prestamos_admin'))

        # Actualizar el préstamo
        conn.execute('''
            UPDATE prestamos 
            SET instructor = ?, jornada = ?, ambiente = ?
            WHERE id = ?
        ''', (instructor, jornada, ambiente, id))
        
        conn.commit()
        
        # Crear notificación
        crear_notificacion(
            'prestamo_editado',
            'Préstamo editado',
            f'Admin editó préstamo de {prestamo["implemento"]} - Instructor: {instructor}',
            session.get('user_id'),
            id
        )
        
        flash(f'Préstamo de {prestamo["implemento"]} actualizado exitosamente', 'success')
        
    except Exception as e:
        flash(f'Error al editar préstamo: {str(e)}', 'error')
    finally:
        conn.close()

    return redirect(url_for('admin.gestion_prestamos_admin'))

# ==================== GESTIÓN DE USUARIOS ====================

# Activar usuario
@admin_bp.route('/activar_usuario/<int:id>', methods=['POST'])
@login_required
def activar_usuario(id):
    if not is_admin():
        return jsonify({'success': False, 'message': 'Sin permisos'}), 403
    
    conn = get_db_connection()
    try:
        # Verificar que el usuario existe
        usuario = conn.execute('SELECT * FROM usuarios WHERE id = ?', (id,)).fetchone()
        if not usuario:
            return jsonify({'success': False, 'message': 'Usuario no encontrado'}), 404
        
        # Activar usuario
        conn.execute('UPDATE usuarios SET activo = 1 WHERE id = ?', (id,))
        conn.commit()
        
        # Crear notificación
        crear_notificacion(
            'usuario_activado',
            'Usuario activado',
            f'El usuario {usuario["nombre"]} ha sido activado por {session.get("user_nombre")}',
            session.get('user_id')
        )
        
        return jsonify({'success': True, 'message': 'Usuario activado exitosamente'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    finally:
        conn.close()

# Desactivar usuario
@admin_bp.route('/desactivar_usuario/<int:id>', methods=['POST'])
@login_required
def desactivar_usuario(id):
    if not is_admin():
        return jsonify({'success': False, 'message': 'Sin permisos'}), 403
    
    conn = get_db_connection()
    try:
        # Verificar que el usuario existe
        usuario = conn.execute('SELECT * FROM usuarios WHERE id = ?', (id,)).fetchone()
        if not usuario:
            return jsonify({'success': False, 'message': 'Usuario no encontrado'}), 404
        
        # No permitir desactivar al admin principal
        if usuario['email'] == 'Eduard@gmail.com':
            return jsonify({'success': False, 'message': 'No se puede desactivar al administrador principal'}), 400
        
        # Desactivar usuario
        conn.execute('UPDATE usuarios SET activo = 0 WHERE id = ?', (id,))
        conn.commit()
        
        # Crear notificación
        crear_notificacion(
            'usuario_desactivado',
            'Usuario desactivado',
            f'El usuario {usuario["nombre"]} ha sido desactivado por {session.get("user_nombre")}',
            session.get('user_id')
        )
        
        return jsonify({'success': True, 'message': 'Usuario desactivado exitosamente'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    finally:
        conn.close()

# Editar usuario
@admin_bp.route('/editar_usuario/<int:id>', methods=['POST'])
@login_required
def editar_usuario(id):
    if not is_admin():
        flash('No tienes permisos para realizar esta acción', 'error')
        return redirect('/admin/usuarios')
    
    conn = get_db_connection()
    try:
        # Obtener datos del formulario
        nombre = request.form.get('nombre')
        email = request.form.get('email')
        telefono = request.form.get('telefono')
        rol = request.form.get('rol')
        
        # Validar campos obligatorios
        if not all([nombre, email, telefono, rol]):
            flash('Todos los campos son obligatorios', 'error')
            return redirect('/admin/usuarios')
        
        # Validar formato de email
        import re
        if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
            flash('Por favor, ingrese un correo electrónico válido', 'error')
            return redirect('/admin/usuarios')
        
        # Verificar que el usuario existe
        usuario_actual = conn.execute('SELECT * FROM usuarios WHERE id = ?', (id,)).fetchone()
        if not usuario_actual:
            flash('Usuario no encontrado', 'error')
            return redirect('/admin/usuarios')
        
        # Verificar si el email ya existe en otro usuario
        email_existente = conn.execute('SELECT id FROM usuarios WHERE email = ? AND id != ?', (email, id)).fetchone()
        if email_existente:
            flash('Este correo electrónico ya está registrado por otro usuario', 'error')
            return redirect('/admin/usuarios')
        
        # Actualizar usuario
        conn.execute('''
            UPDATE usuarios 
            SET nombre = ?, email = ?, telefono = ?, rol = ?
            WHERE id = ?
        ''', (nombre, email, telefono, rol, id))
        
        conn.commit()
        
        # Crear notificación
        crear_notificacion(
            'usuario_editado',
            'Usuario editado',
            f'El usuario {nombre} ha sido editado por {session.get("user_nombre")}',
            session.get('user_id')
        )
        
        flash(f'Usuario {nombre} actualizado exitosamente', 'success')
        
    except Exception as e:
        flash(f'Error al actualizar usuario: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect('/admin/usuarios')

# Eliminar usuario
@admin_bp.route('/eliminar_usuario/<int:id>', methods=['POST'])
@login_required
def eliminar_usuario(id):
    if not is_admin():
        return jsonify({'success': False, 'message': 'Sin permisos'}), 403
    
    conn = get_db_connection()
    try:
        # Verificar que el usuario existe
        usuario = conn.execute('SELECT * FROM usuarios WHERE id = ?', (id,)).fetchone()
        if not usuario:
            return jsonify({'success': False, 'message': 'Usuario no encontrado'}), 404
        
        # No permitir eliminar al admin principal
        if usuario['email'] == 'Eduard@gmail.com':
            return jsonify({'success': False, 'message': 'No se puede eliminar al administrador principal'}), 400
        
        # Verificar si tiene préstamos activos
        prestamos_activos = conn.execute(
            'SELECT COUNT(*) as count FROM prestamos WHERE fk_usuario = ? AND fecha_devolucion IS NULL',
            (id,)
        ).fetchone()
        
        if prestamos_activos['count'] > 0:
            return jsonify({'success': False, 'message': 'No se puede eliminar un usuario con préstamos activos'}), 400
        
        # Eliminar usuario
        conn.execute('DELETE FROM usuarios WHERE id = ?', (id,))
        conn.commit()
        
        # Crear notificación
        crear_notificacion(
            'usuario_eliminado',
            'Usuario eliminado',
            f'El usuario {usuario["nombre"]} ha sido eliminado por {session.get("user_nombre")}',
            session.get('user_id')
        )
        
        return jsonify({'success': True, 'message': 'Usuario eliminado exitosamente'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    finally:
        conn.close()

# API para obtener usuarios pendientes
@admin_bp.route('/api/usuarios_pendientes')
@login_required
def api_usuarios_pendientes():
    if not is_admin():
        return jsonify({'error': 'Sin permisos'}), 403
    
    conn = get_db_connection()
    try:
        usuarios_pendientes = conn.execute('''
            SELECT id, nombre, email, rol, fecha_registro
            FROM usuarios 
            WHERE activo = 0 
            ORDER BY fecha_registro DESC
        ''').fetchall()
        
        return jsonify([dict(usuario) for usuario in usuarios_pendientes])
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

# ==================== REPORTES ====================

# Página de reportes
@admin_bp.route('/reportes')
@login_required
def reportes():
    if not is_admin():
        flash('No tienes permisos para acceder a esta página', 'error')
        return redirect('/')
    
    conn = get_db_connection()
    try:
        # Obtener estadísticas básicas para el dashboard
        total_prestamos = conn.execute('SELECT COUNT(*) as count FROM prestamos').fetchone()['count']
        prestamos_activos = conn.execute('SELECT COUNT(*) as count FROM prestamos WHERE fecha_devolucion IS NULL').fetchone()['count']
        prestamos_devueltos = conn.execute('SELECT COUNT(*) as count FROM prestamos WHERE fecha_devolucion IS NOT NULL').fetchone()['count']
        
        # Obtener implementos más prestados
        implementos_mas_prestados = conn.execute('''
            SELECT i.implemento, COUNT(p.id) as total_prestamos
            FROM implementos i
            LEFT JOIN prestamos p ON i.id = p.fk_implemento
            GROUP BY i.id, i.implemento
            ORDER BY total_prestamos DESC
            LIMIT 10
        ''').fetchall()
        
        # Obtener usuarios más activos
        usuarios_activos = conn.execute('''
            SELECT u.nombre, u.email, COUNT(p.id) as total_prestamos
            FROM usuarios u
            LEFT JOIN prestamos p ON u.id = p.fk_usuario
            GROUP BY u.id, u.nombre, u.email
            ORDER BY total_prestamos DESC
            LIMIT 10
        ''').fetchall()
        
        conn.close()
        
        return render_template('admin/reportes.html',
                             total_prestamos=total_prestamos,
                             prestamos_activos=prestamos_activos,
                             prestamos_devueltos=prestamos_devueltos,
                             implementos_mas_prestados=implementos_mas_prestados,
                             usuarios_activos=usuarios_activos)
    except Exception as e:
        flash(f'Error al cargar reportes: {str(e)}', 'error')
        conn.close()
        return redirect('/admin')

# Generar reporte de préstamos en Excel
@admin_bp.route('/reportes/prestamos/excel', methods=['POST'])
@login_required
def generar_reporte_prestamos_excel():
    if not is_admin():
        return jsonify({'error': 'Sin permisos'}), 403
    
    try:
        # Obtener parámetros del formulario
        fecha_inicio = request.form.get('fecha_inicio')
        fecha_fin = request.form.get('fecha_fin')
        tipo_reporte = request.form.get('tipo_reporte', 'todos')  # todos, activos, devueltos
        formato = request.form.get('formato', 'detallado')  # detallado, resumen
        
        print(f"DEBUG: Generando reporte - Tipo: {tipo_reporte}, Formato: {formato}")
        print(f"DEBUG: Fechas - Inicio: {fecha_inicio}, Fin: {fecha_fin}")
        
        # Validar fechas
        if fecha_inicio and fecha_fin:
            try:
                from datetime import datetime
                datetime.strptime(fecha_inicio, '%Y-%m-%d')
                datetime.strptime(fecha_fin, '%Y-%m-%d')
            except ValueError:
                return jsonify({'error': 'Formato de fecha inválido'}), 400
        
        # Construir query
        conn = get_db_connection()
        
        query = '''
            SELECT 
                p.id,
                p.tipo_prestamo,
                p.nombre_prestatario,
                p.ficha,
                p.ambiente,
                p.horario,
                p.instructor,
                p.jornada,
                p.fecha_prestamo,
                p.fecha_devolucion,
                p.novedad,
                p.estado_implemento_devolucion,
                p.observaciones,
                i.implemento,
                i.categoria,
                u.nombre as usuario_registro,
                u.email as usuario_email
            FROM prestamos p
            JOIN implementos i ON p.fk_implemento = i.id
            JOIN usuarios u ON p.fk_usuario = u.id
            WHERE 1=1
        '''
        
        params = []
        
        # Filtros de fecha
        if fecha_inicio:
            query += " AND DATE(p.fecha_prestamo) >= ?"
            params.append(fecha_inicio)
        
        if fecha_fin:
            query += " AND DATE(p.fecha_prestamo) <= ?"
            params.append(fecha_fin)
        
        # Filtro de tipo
        if tipo_reporte == 'activos':
            query += " AND p.fecha_devolucion IS NULL"
        elif tipo_reporte == 'devueltos':
            query += " AND p.fecha_devolucion IS NOT NULL"
        
        query += " ORDER BY p.fecha_prestamo DESC"
        
        prestamos_raw = conn.execute(query, params).fetchall()
        conn.close()
        
        # Convertir Row objects a diccionarios
        prestamos = [dict(prestamo) for prestamo in prestamos_raw]
        print(f"DEBUG: {len(prestamos)} préstamos encontrados para el reporte")
        
        # Generar Excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        from datetime import datetime
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Préstamos"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Configurar columnas según el formato
        if formato == 'detallado':
            headers = [
                'ID', 'Tipo Préstamo', 'Prestatario', 'Ficha', 'Ambiente', 'Horario',
                'Instructor', 'Jornada', 'Implemento', 'Categoría', 'Fecha Préstamo',
                'Fecha Devolución', 'Estado', 'Novedad', 'Observaciones', 'Registrado por'
            ]
            columns = [
                'id', 'tipo_prestamo', 'nombre_prestatario', 'ficha', 'ambiente', 'horario',
                'instructor', 'jornada', 'implemento', 'categoria', 'fecha_prestamo',
                'fecha_devolucion', 'estado_implemento_devolucion', 'novedad', 'observaciones', 'usuario_registro'
            ]
        else:  # resumen
            headers = [
                'ID', 'Prestatario', 'Implemento', 'Fecha Préstamo', 'Fecha Devolución',
                'Días Transcurridos', 'Estado', 'Registrado por'
            ]
            columns = [
                'id', 'nombre_prestatario', 'implemento', 'fecha_prestamo',
                'fecha_devolucion', 'dias_transcurridos', 'estado', 'usuario_registro'
            ]
        
        # Escribir encabezados
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Escribir datos
        for row, prestamo in enumerate(prestamos, 2):
            for col, column_name in enumerate(columns, 1):
                if column_name == 'dias_transcurridos':
                    # Calcular días transcurridos
                    if prestamo.get('fecha_devolucion'):
                        fecha_fin = datetime.strptime(prestamo['fecha_devolucion'][:10], '%Y-%m-%d')
                        fecha_inicio = datetime.strptime(prestamo['fecha_prestamo'][:10], '%Y-%m-%d')
                        dias = (fecha_fin - fecha_inicio).days
                        value = f"{dias} días"
                    else:
                        fecha_inicio = datetime.strptime(prestamo['fecha_prestamo'][:10], '%Y-%m-%d')
                        dias = (datetime.now() - fecha_inicio).days
                        value = f"{dias} días (activo)"
                elif column_name == 'estado':
                    if prestamo.get('fecha_devolucion'):
                        value = 'Devuelto'
                    else:
                        value = 'Activo'
                else:
                    # Manejar valores nulos de forma segura
                    raw_value = prestamo.get(column_name)
                    if raw_value is None or raw_value == '':
                        value = ''
                    else:
                        value = str(raw_value)
                
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical='center')
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15
        
        # Agregar información del reporte
        ws.cell(row=len(prestamos) + 3, column=1, value="Información del Reporte:")
        ws.cell(row=len(prestamos) + 4, column=1, value=f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=len(prestamos) + 5, column=1, value=f"Total de registros: {len(prestamos)}")
        ws.cell(row=len(prestamos) + 6, column=1, value=f"Tipo de reporte: {tipo_reporte}")
        ws.cell(row=len(prestamos) + 7, column=1, value=f"Período: {fecha_inicio or 'Inicio'} - {fecha_fin or 'Actual'}")
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre del archivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"reporte_prestamos_{tipo_reporte}_{timestamp}.xlsx"
        
        from flask import make_response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        print(f"DEBUG: Reporte generado exitosamente - {len(prestamos)} registros")
        return response
        
    except Exception as e:
        print(f"ERROR: Error al generar reporte: {e}")
        import traceback
        traceback.print_exc()
        
        # Mensaje de error más amigable
        error_message = str(e)
        if "'sqlite3.Row' object has no attribute 'get'" in error_message:
            error_message = "Error interno del sistema. Por favor, contacte al administrador."
        elif "openpyxl" in error_message.lower():
            error_message = "Error al generar el archivo Excel. Verifique que la librería esté instalada correctamente."
        
        return jsonify({'error': f'Error al generar reporte: {error_message}'}), 500
        
# API para obtener instructores disponibles
@admin_bp.route('/api/instructores_disponibles')
@login_required
def api_instructores_disponibles():
    conn = get_db_connection()
    try:
        instructores = conn.execute('''
            SELECT id, nombre, email
            FROM usuarios 
            WHERE rol = 'instructor' AND activo = 1
            ORDER BY nombre ASC
        ''').fetchall()
        
        return jsonify([dict(instructor) for instructor in instructores])
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()